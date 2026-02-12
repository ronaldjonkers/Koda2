"""Expense management service with receipt processing."""

from __future__ import annotations

import base64
import datetime as dt
from decimal import Decimal
from pathlib import Path
from typing import Any, Optional

from koda2.config import get_settings
from koda2.logging_config import get_logger
from koda2.modules.expenses.models import Expense, ExpenseCategory, ExpenseReport, ExpenseStatus
from koda2.modules.llm import LLMRouter

logger = get_logger(__name__)


class ExpenseService:
    """Service for expense management and receipt processing."""
    
    def __init__(self, llm_router: Optional[LLMRouter] = None) -> None:
        self._settings = get_settings()
        self._llm = llm_router
        self._reports: dict[str, ExpenseReport] = {}
        
    def set_llm_router(self, router: LLMRouter) -> None:
        """Set LLM router for receipt processing."""
        self._llm = router
        
    async def process_receipt(
        self,
        image_path: str,
        submitted_by: str,
        project_code: Optional[str] = None,
    ) -> Expense:
        """Process a receipt image and extract expense data using GPT-4 Vision."""
        
        # Read image
        path = Path(image_path)
        if not path.exists():
            raise FileNotFoundError(f"Receipt image not found: {image_path}")
            
        image_bytes = path.read_bytes()
        base64_image = base64.b64encode(image_bytes).decode("utf-8")
        
        # Determine MIME type
        mime_type = "image/jpeg"
        if path.suffix.lower() == ".png":
            mime_type = "image/png"
        elif path.suffix.lower() == ".pdf":
            mime_type = "application/pdf"
            
        # Use GPT-4 Vision to extract data
        if self._llm and hasattr(self._llm, '_providers'):
            try:
                expense_data = await self._extract_receipt_data(base64_image, mime_type)
            except Exception as e:
                logger.error("receipt_extraction_failed", error=str(e))
                expense_data = {}
        else:
            expense_data = {}
            
        # Create expense
        expense = Expense(
            date=expense_data.get("date", dt.date.today()),
            description=expense_data.get("description", "Unknown expense"),
            category=expense_data.get("category", ExpenseCategory.OTHER),
            amount=Decimal(str(expense_data.get("amount", 0))),
            currency=expense_data.get("currency", "EUR"),
            vat_amount=expense_data.get("vat_amount"),
            vat_percentage=expense_data.get("vat_percentage"),
            merchant_name=expense_data.get("merchant_name"),
            merchant_address=expense_data.get("merchant_address"),
            merchant_vat=expense_data.get("merchant_vat"),
            receipt_image_path=image_path,
            receipt_extracted_text=expense_data.get("raw_text", ""),
            project_code=project_code,
            submitted_by=submitted_by,
        )
        
        logger.info("receipt_processed", 
                   expense_id=expense.id,
                   amount=str(expense.amount),
                   merchant=expense.merchant_name)
                   
        return expense
        
    async def _extract_receipt_data(
        self, 
        base64_image: str, 
        mime_type: str
    ) -> dict[str, Any]:
        """Extract receipt data using GPT-4 Vision."""
        import openai
        
        client = openai.AsyncOpenAI(api_key=self._settings.openai_api_key)
        
        prompt = """Analyze this receipt and extract the following information:
        
        1. Date of purchase (YYYY-MM-DD format)
        2. Merchant/Store name
        3. Total amount (numeric only)
        4. Currency (EUR, USD, etc.)
        5. VAT/BTW amount if visible
        6. VAT percentage if visible (21%, 9%, etc.)
        7. Merchant address
        8. Merchant VAT/BTW number if visible
        9. Brief description of items purchased
        10. Category (choose from: travel, accommodation, meals, transport, office_supplies, entertainment, other)
        11. Full extracted text for reference
        
        Respond in JSON format:
        {
            "date": "2024-01-15",
            "merchant_name": "Store Name",
            "amount": 125.50,
            "currency": "EUR",
            "vat_amount": 21.80,
            "vat_percentage": 21.0,
            "merchant_address": "Street 1, City",
            "merchant_vat": "NL123456789B01",
            "description": "Office supplies",
            "category": "office_supplies",
            "raw_text": "full receipt text"
        }
        
        If any field is not visible, use null."""
        
        response = await client.chat.completions.create(
            model="gpt-4o",
            messages=[
                {
                    "role": "user",
                    "content": [
                        {"type": "text", "text": prompt},
                        {
                            "type": "image_url",
                            "image_url": {
                                "url": f"data:{mime_type};base64,{base64_image}"
                            }
                        }
                    ]
                }
            ],
            max_tokens=1000,
        )
        
        content = response.choices[0].message.content or "{}"
        
        # Parse JSON
        import json
        try:
            # Extract JSON from code blocks if present
            if "```json" in content:
                content = content.split("```json")[1].split("```")[0]
            elif "```" in content:
                content = content.split("```")[1].split("```")[0]
                
            data = json.loads(content.strip())
            
            # Convert category string to enum
            category_map = {
                "travel": ExpenseCategory.TRAVEL,
                "accommodation": ExpenseCategory.ACCOMMODATION,
                "meals": ExpenseCategory.MEALS,
                "transport": ExpenseCategory.TRANSPORT,
                "office_supplies": ExpenseCategory.OFFICE_SUPPLIES,
                "entertainment": ExpenseCategory.ENTERTAINMENT,
                "training": ExpenseCategory.TRAINING,
                "communication": ExpenseCategory.COMMUNICATION,
                "other": ExpenseCategory.OTHER,
            }
            
            if "category" in data and isinstance(data["category"], str):
                data["category"] = category_map.get(data["category"].lower(), ExpenseCategory.OTHER)
                
            return data
            
        except json.JSONDecodeError as e:
            logger.error("receipt_json_parse_failed", content=content, error=str(e))
            return {}
            
    async def create_report(
        self,
        title: str,
        employee_name: str,
        period_start: dt.date,
        period_end: dt.date,
        description: str = "",
        department: Optional[str] = None,
    ) -> ExpenseReport:
        """Create a new expense report."""
        report = ExpenseReport(
            title=title,
            description=description,
            employee_name=employee_name,
            department=department,
            period_start=period_start,
            period_end=period_end,
        )
        self._reports[report.id] = report
        logger.info("expense_report_created", report_id=report.id, employee=employee_name)
        return report
        
    async def export_to_excel(self, report: ExpenseReport) -> str:
        """Export expense report to Excel."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        
        output_dir = Path("data/expenses")
        output_dir.mkdir(parents=True, exist_ok=True)
        output_path = output_dir / f"expense_report_{report.id}.xlsx"
        
        wb = Workbook()
        
        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        # Header
        ws_summary["A1"] = "Expense Report"
        ws_summary["A1"].font = Font(size=16, bold=True)
        ws_summary.merge_cells("A1:E1")
        
        ws_summary["A3"] = "Employee:"
        ws_summary["B3"] = report.employee_name
        ws_summary["A4"] = "Department:"
        ws_summary["B4"] = report.department or "N/A"
        ws_summary["A5"] = "Period:"
        ws_summary["B5"] = f"{report.period_start} to {report.period_end}"
        ws_summary["A6"] = "Status:"
        ws_summary["B6"] = report.status.value
        
        # Totals
        ws_summary["A8"] = "Total Amount:"
        ws_summary["B8"] = float(report.total_amount)
        ws_summary["B8"].number_format = '#,##0.00'
        ws_summary["A9"] = "Total VAT:"
        ws_summary["B9"] = float(report.total_vat)
        ws_summary["B9"].number_format = '#,##0.00'
        
        # Category breakdown
        ws_summary["A11"] = "Category Breakdown"
        ws_summary["A11"].font = Font(bold=True)
        
        ws_summary["A12"] = "Category"
        ws_summary["B12"] = "Amount"
        ws_summary["A12"].font = Font(bold=True)
        ws_summary["B12"].font = Font(bold=True)
        
        row = 13
        for category, total in report.get_category_totals().items():
            ws_summary[f"A{row}"] = category.value
            ws_summary[f"B{row}"] = float(total)
            ws_summary[f"B{row}"].number_format = '#,##0.00'
            row += 1
            
        # Expenses sheet
        ws_expenses = wb.create_sheet("Expenses")
        
        headers = ["Date", "Category", "Description", "Merchant", "Amount", "VAT", "Currency", "Project", "Status"]
        for col, header in enumerate(headers, 1):
            cell = ws_expenses.cell(1, col, header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
        for row, expense in enumerate(report.expenses, 2):
            ws_expenses.cell(row, 1, expense.date)
            ws_expenses.cell(row, 2, expense.category.value)
            ws_expenses.cell(row, 3, expense.description)
            ws_expenses.cell(row, 4, expense.merchant_name or "")
            ws_expenses.cell(row, 5, float(expense.amount))
            ws_expenses.cell(row, 5).number_format = '#,##0.00'
            ws_expenses.cell(row, 6, float(expense.calculate_vat()))
            ws_expenses.cell(row, 6).number_format = '#,##0.00'
            ws_expenses.cell(row, 7, expense.currency)
            ws_expenses.cell(row, 8, expense.project_code or "")
            ws_expenses.cell(row, 9, expense.status.value)
            
        # Adjust column widths
        for ws in [ws_summary, ws_expenses]:
            for col in range(1, 10):
                ws.column_dimensions[get_column_letter(col)].width = 15
                
        wb.save(str(output_path))
        report.excel_path = str(output_path)
        
        logger.info("expense_report_exported", report_id=report.id, path=str(output_path))
        return str(output_path)
        
    async def scan_email_for_receipts(
        self,
        email_service: Any,
        since: Optional[dt.datetime] = None,
    ) -> list[Expense]:
        """Scan emails for receipt attachments and process them."""
        expenses = []
        
        # Search for emails with receipt/bon/invoice keywords
        keywords = ["receipt", "invoice", "bon", "factuur", "payment confirmation"]
        
        # This would integrate with the email service to search attachments
        # For now, return empty list as placeholder
        logger.info("email_receipt_scan_started", keywords=keywords)
        
        return expenses
        
    def get_report(self, report_id: str) -> Optional[ExpenseReport]:
        """Get an expense report by ID."""
        return self._reports.get(report_id)
        
    def get_employee_reports(self, employee_name: str) -> list[ExpenseReport]:
        """Get all expense reports for an employee."""
        return [
            r for r in self._reports.values()
            if r.employee_name == employee_name
        ]
        
    def get_pending_reports(self) -> list[ExpenseReport]:
        """Get all pending expense reports."""
        return [
            r for r in self._reports.values()
            if r.status == ExpenseStatus.PENDING
        ]
